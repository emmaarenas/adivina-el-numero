# Importo todo lo que voy a utilizar
import random # Para generar números aleatorios
import Adivinanza as A # Donde se ubican mis funciones para el juego de adivinanza
from openpyxl import load_workbook # Para cargar el Excel con las estadísticas

# Creo una función para el menú del juego
def menu():
    while True: # Bucle para que el usuario pueda elegir la opción deseada
        print("\n\033[1;30;47mMenú:\033[0m")
        try: # Por si la opcion se introduce como un valor no numérico
            opcion=int(input("\033[1m1. Partida modo solitario\n2. Partida 2 jugadores\n3. Estadística\n4. Salir\n\033[0m"))
            if opcion==1: # Si ha elegido la opcion solitario
                solitario(opcion) # Se llama a la función solitario
            elif opcion==2: # Si ha elegido la opcion 2 jugadores
                multijugador(opcion) # Se llama a la función multijugador
            elif opcion==3: # Si ha elegido la opción estadística
                estadistica() # Se llama a la función estadística
            elif opcion==4: # Si ha elegido salir del juego
                print("\n\033[1;33;47m¡Gracias por jugar! Hasta luego.\033[0m")
                exit() # Salimos del programa
            else: # Si se introduce un número diferente a 1,2,3 o 4
                print("\033[4mEsa opción no es válida.\033[0m")
        except ValueError: # Si se introduce un valor no numérico
            print("\033[4mPor favor, ingresa un número válido.\033[0m")

# Creo una función para la opcion de juego solitario
def solitario(opcion):
    print("\n\033[1;33;47m¡Has escogido el modo Solitario!\033[0m")
    dificultad=A.elegir_dificultad() # Llamo a la función elegir_dificultad de mi módulo Adivinanza
    # Configuro los parámetros necesario
    intentos={1:20,2:12,3:5}[dificultad] # Uso un diccionario para almacenar los intentos por dificultad
    maximo={1:1000,2:2000,3:3000}[dificultad] # Uso un diccionario para almacenar los maximos por dificultad                    
    numero_aleatorio=random.randint(1, maximo) # Se genera un número aleatorio entre 1 y el máximo según la dificultad
    no_acertado=[] # Una lista está vacía para almacenar los números ya introducidos
    intentos_restantes=intentos # El número de intentos restantes es igual al número de intentos por dificultad  
    print(f"\n\033[1;3mPista: El número a adivinar se encuentra entre 1 y {maximo}\033[0m") 
    # Creo un bucle para adivinar el número (con el número de intentos según la dificultad escogida)
    A.adivinar(intentos,intentos_restantes,maximo,no_acertado,numero_aleatorio,dificultad,opcion) # Llamo a la función adivinar de mi módulo Adivinanza

# Creo una función para la opcion de 2 jugadores
def multijugador(opcion):
    print("\n\033[1;33;47m¡Has escogido el modo 2 Jugadores!\033[0m")
    dificultad=A.elegir_dificultad() # Llamo a la función elegir_dificultad de mi módulo Adivinanza
    # Configuro los parámetros necesario
    intentos={1:20,2:12,3:5}[dificultad] # Uso un diccionario para almacenar los intentos por dificultad
    maximo={1:1000,2:2000,3:3000}[dificultad] # Uso un diccionario para almacenar los maximos por dificultad
    no_acertado=[] # Una lista está vacía para almacenar los números ya introducidos
    intentos_restantes=intentos # El número de intentos restantes es igual al número de intentos por dificultad
    print("\n\033[1;37;42mJugador 1: Es tu turno.\033[0m")
    # Llamo a la función iintrodur_numero para que el Jugador 1 introduzca el número  
    numero_aleatorio=A.introducir_numero(maximo) # Asigno a numero_aleatorio el número introducido por el usuario
    print("\n\033[1;37;42mJugador 2: Es tu turno.\033[0m")
    print(f"\n\033[1;3mPista: El número a adivinar se encuentra entre 1 y {maximo}\033[0m")    
    # Creo un bucle para adivinar el número (con el número de intentos según la dificultad escogida)
    A.adivinar(intentos,intentos_restantes,maximo,no_acertado,numero_aleatorio,dificultad,opcion) # Llamo a la función adivinar de mi módulo Adivinanza

# Creo una función para las estadísticas
def estadistica():
    try: # Intento cargar el archivo Excel existente
        workbook=load_workbook("./Estadistica.xlsx")
        hoja=workbook["Estadísticas"]
        if hoja.max_row > 1:  # Si hay más de una fila, significa que hay datos guardados
            print("\n\033[1;33;47mEstadísticas del Juego\033[0m\n") 
            # Muestro las cabeceras
            cabeceras=["Fecha","Jugador","Resultado","Dificultad","Modo","Intentos"]
            print(f"{cabeceras[0]:<20}{cabeceras[1]:<12}{cabeceras[2]:<10}{cabeceras[3]:<12}{cabeceras[4]:<15}{cabeceras[5]:<10}")
            print("-"*82) # Muestro líneas para simular una tabla
            # Recorro las filas y muestro cada dato
            for fila in hoja.iter_rows(min_row=2,values_only=True):
                fecha,jugador,resultado,dificultad,modo_juego,intentos_usados=fila
                if dificultad==1:
                    dificultad_str="Fácil" 
                elif dificultad==2:
                    dificultad_str="Medio"
                else:
                    dificultad_str="Difícil"
                print(f"{fecha:<20}{jugador:<12}{resultado:<10}{dificultad_str:<12}{modo_juego:<15}{intentos_usados:<10}")
            while True: # Bucle para que el usuario pueda elegir la opción deseada
                print("\n\033[1;30;47m¿Deseas visualizar algún gráfico?\033[0m")
                try: # Por si la opcion se introduce como un valor no numérico
                    grafico=int(input("\033[1m1. Dificultades por partidas\n2. Resultados por modo\n3. Intentos por dificultad\n4. Volver al menú principal\n\033[0m"))
                    if grafico==1: # Si ha elegido el gráfico primero
                        A.grafico_dificultades()# Se llama a la función correspondiente
                    elif grafico==2: # Si ha elegido el gráfico segundo
                        A.grafico_resultados() # Se llama a la función correspondiente
                    elif grafico==3: # Si ha elegido el gráfico tercero
                        A.grafico_intentos() # Se llama a la función correspondiente
                    elif grafico==4: # Si ha elegido volver al menú principal
                        return # Volvemos al menú
                    else: # Si se introduce un número diferente a 1,2,3 o 4
                        print("\033[4mEsa opción no es válida.\033[0m")
                except ValueError: # Si se introduce un valor no numérico
                    print("\033[4mPor favor, ingresa un número válido.\033[0m")
        else:
            print("\033[4mNo existen datos guardados actualmente.\033[0m")
    except FileNotFoundError: # Si no hay Excel creado
        print("\033[4mNo existen datos guardados actualmente.\033[0m")
        return # Volvemos al menú
    
# Inicio el juego
print("\n\033[1;33;47m¡Bienvenido a Adivina el Número!\033[0m")
menu()