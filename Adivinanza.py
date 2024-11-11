# Importo todo lo que voy a utilizar
import getpass # Para ocultar la entrada introducida
import openpyxl # Para modificar el Excel con las estadísticas
from datetime import datetime
from openpyxl import load_workbook # Para cargar el Excel con las estadísticas
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from collections import Counter # Para contar ocurrencias
import matplotlib.pyplot as plt # Para generar el gráfico con las estadísticas

# Creo una función para elegir la dificultad del juego 
def elegir_dificultad():
    while True: # Creo un bucle para elegir la dificultad
        try: # Por si la dificultad se introduce como un valor no numérico
            dificultad=int(input("\n\033[1;30;47mElige la dificultad\033[0m\033[1m\n1. Fácil\n2. Medio\n3. Difícil\n\033[0m"))
            if (1<=dificultad<=3): # Si la dificultad está en el rango permitido
                return dificultad
            else: # Si la dificultad no está en el rango permitido
                print("\033[4mLa dificultad debe estar entre 1 y 3.\033[0m")
        except ValueError: # Si se introduce un valor no numérico
            print("\033[4mPor favor, ingresa un número válido.\033[0m")

# Creo una función para dar la pista
def pista(numero_introducido,numero_aleatorio):
    if (numero_introducido<numero_aleatorio): # Si el número introducido es menor que el número a adivinar
        print(f"\033[1;3mPista: El número {numero_introducido} es menor que el número a adivinar\033[0m")
    else: # Si el número introducido es mayor que el número a adivinar
        print(f"\033[1;3mPista: El número {numero_introducido} es mayor que el número a adivinar\033[0m")

# Creo una función para introducir el número a adivinar en el modo 2 jugadores
def introducir_numero(maximo):
    while True:
        try: # Por si el número se introduce como un valor no numérico
            numero_aleatorio=int(getpass.getpass(f"\n\033[1mIntroduce un número entre 1 y {maximo}: \033[0m")) # Uso de getpass para ocultar el número introducido
            if (numero_aleatorio>maximo or numero_aleatorio<1):
                print("\033[4mEl número está fuera del rango\033[0m") 
            else:
                return numero_aleatorio
        except ValueError: # Si se introduce un valor no numérico
            print("\033[4mPor favor, ingresa un número válido.\033[0m")

# Creo una función para adivinar el número
def adivinar(intentos,intentos_restantes,maximo,no_acertado,numero_aleatorio,dificultad,opcion):
    while intentos_restantes>0: # He elegido el bucle while en vez de for para que solo se reste el intento cuando el número introducido no sea el correcto y tampoco esté en la lista de no_acertados
        try: # Por si el número se introduce como un valor no numérico
            numero_introducido=int(input(f"\033[1m\nTrata de adivinarlo, tienes {intentos_restantes} intentos: \033[0m")) # Le muestro el número de intentos restantes que le quedan
        except ValueError: # Si se introduce un valor no numérico
            print("\033[4mPor favor, ingresa un número válido.\033[0m")
            continue # Paso al siguiente intento
        if (numero_introducido<1 or numero_introducido>maximo): # Si el número no está en el rango permitido    
            print(f"\033[4mEl número debe estar entre 1 y {maximo}.\033[0m")
            continue # Paso al siguiente intento
        if numero_introducido in no_acertado: # Si el número ya lo ha probado antes
            print("\033[4mEse número ya lo has probado...\033[0m")
            continue # Paso al siguiente intento
        if numero_introducido==numero_aleatorio: # Si ha acertado el número
            print("\n\033[1;33;47m¡Enhorabuena! Has acertado el número.\033[0m")
            ganado=True
            guardar(intentos,intentos_restantes,ganado,dificultad,opcion)
            return # Volvemos al menú
        else:
            no_acertado.append(numero_introducido) # Añado el número a la lista de los intentos fallidos
            intentos_restantes-=1 # Decremento el número de intentos restantes
            print(f"\033[1;37;44m¡Que pena, no lo has adivinado!\033[0m")
            pista(numero_introducido,numero_aleatorio)
    else: # Si se ha agotado el número de intentos
        print(f"\n\033[1;37;44m¡Has perdido! El número era {numero_aleatorio}\033[0m") # Le muestro el número que debía adivinar
        ganado=False
        guardar(intentos,intentos_restantes,ganado,dificultad,opcion)

# Creo una función para crear el Excel con los datos
def crear_excel():
    try: # Intento abrir el archivo si es que existe
        workbook=openpyxl.load_workbook("./Estadistica.xlsx")
    except FileNotFoundError: # Si no existe, creo uno nuevo
        workbook=openpyxl.Workbook()
        hoja=workbook.active # Selecciono la hoja activa (la que se crea por defecto)
        hoja.title="Estadísticas" # Le cambio el nombre a esa hoja
        cabeceras=["Fecha","Jugador","Resultado","Dificultad","Modo","Intentos"] # Creo las cabeceras
        # Añado y aplico un formato bonito a las cabeceras
        for col, cabecera in enumerate(cabeceras, start=1):
            cell=hoja.cell(row=1, column=col, value=cabecera) 
            cell.font=Font(bold=True, color="FFFFFF") 
            hoja.column_dimensions[get_column_letter(col)].width=20
            cell.fill=openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        hoja.auto_filter.ref="A1:F1" # Añado los filtros ya de antemano
        workbook.save("./Estadistica.xlsx") # Guardo el Excel

# Creo una función para guardar los datos de la partida
def guardar(intentos,intentos_restantes,ganado,dificultad,opcion):
    nombre_jugador=str(input("\n\033[1mPor favor, escribe tu nombre: \033[0m")) # Se le pide el nombre del jugador
    fecha_actual=datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Esto lo he encontrado en internet, para guardar la fecha de la partida
    # Asigno el valor del resultado y los intentos en función de si ha ganado o no
    if ganado==True: 
        resultado="Ganado"
    elif ganado==False:
        resultado="Perdido"
    # Asigno el modo de juego según la opción que escogieron
    if opcion==1:
        modo_juego="Solitario"
    elif opcion==2:
        modo_juego="2 jugadores"
    # Calculo los intentos usados en las partidas
    intentos_usados=intentos-intentos_restantes
    crear_excel() # Llamo a la función crear_excel
    workbook=load_workbook("./Estadistica.xlsx") # Abro el archivo
    hoja=workbook["Estadísticas"] # Selecciono la hoja
    siguiente_fila=hoja.max_row+1 # Busco la siguiente fila vacía para añadirlos ahí
    # Agrego los datos del juego
    hoja.cell(row=siguiente_fila,column=1,value=fecha_actual)
    hoja.cell(row=siguiente_fila,column=2,value=nombre_jugador)
    hoja.cell(row=siguiente_fila,column=3,value=resultado)
    hoja.cell(row=siguiente_fila,column=4,value=dificultad)
    hoja.cell(row=siguiente_fila,column=5,value=modo_juego)
    hoja.cell(row=siguiente_fila,column=6,value=intentos_usados)
    workbook.save("./Estadistica.xlsx") # Guardo los cambios

# Creo una función para generar el primer gráfico
def grafico_dificultades():
    # Creo las listas vacías para generar el gráfico 
    dificultades=[]
    colores={"Fácil":"green","Medio":"orange","Difícil":"red"} # Asigno colores a cada dificultad
    # Abro el archivo Excel para leer los datos
    workbook=load_workbook("./Estadistica.xlsx")
    hoja=workbook["Estadísticas"]
    # Recorro las filas y agrego cada dificultad a la lista
    for fila in hoja.iter_rows(min_row=2,values_only=True):
        _, _, _, dificultad, _, _ = fila
        if dificultad == 1:
            dificultades.append("Fácil")
        elif dificultad == 2:
            dificultades.append("Medio")
        elif dificultad == 3:
            dificultades.append("Difícil")
    # ContabilizO las ocurrencias de cada dificultad
    dificultad_counts=Counter(dificultades)
    # Genero el gráfico de barras para la dificultad
    plt.figure(figsize=(10,5))  # Configuro el tamaño del gráfico
    plt.bar(dificultad_counts.keys(),dificultad_counts.values(),
            color=[colores[dif] for dif in dificultad_counts.keys()],edgecolor="black")
    plt.title("Distribución de dificultades en el juego")  # Título del gráfico
    plt.xlabel("Dificultad")  # Etiqueta del eje X
    plt.ylabel("Número de partidas")  # Etiqueta del eje Y
    plt.show()
    return

# Creo una función para generar el segundo gráfico
def grafico_resultados():
    # Creo unos contadores para los modos y resultados
    resultados_modos={
        "Solitario":{"Ganado":0,"Perdido": 0},
        "2 jugadores":{"Ganado":0,"Perdido": 0}
    }
    # Abro el archivo Excel para leer los datos
    workbook=load_workbook("./Estadistica.xlsx")
    hoja=workbook["Estadísticas"]
    # Recorro las filas y cuento los resultados por modo
    for fila in hoja.iter_rows(min_row=2,values_only=True):
        _, _, resultado, _, modo_juego, _ = fila
        if modo_juego in resultados_modos:
            if resultado in resultados_modos[modo_juego]:
                resultados_modos[modo_juego][resultado]+=1
    # Preparo los datos para crear el gráfico
    modos=list(resultados_modos.keys())
    ganados=[resultados_modos[modo]["Ganado"] for modo in modos]
    perdidos=[resultados_modos[modo]["Perdido"] for modo in modos]
    # Configuro el gráfico
    x=range(len(modos))  # Posiciones para cada modo en el eje x
    plt.figure(figsize=(10,5))
    plt.bar(x,ganados,width=0.4,label="Ganado",color="blue",align='center')
    plt.bar(x,perdidos,width=0.4,label="Perdido",color="red",align='edge')
    # Configuro etiquetas y leyendas
    plt.title("Resultados de partidas por modo de juego")
    plt.xlabel("Modo de juego")
    plt.ylabel("Número de partidas")
    plt.xticks(x,modos)  # Asigno los modos al eje x
    plt.legend()
    plt.show() # Muestro el gráfico
    return

# Creo una función para generar el tercer gráfico
def grafico_intentos():
    # Diccionario para almacenar la suma de intentos y el conteo de partidas por dificultad
    intentos_dificultad={"Fácil": {"intentos":0,"partidas":0},
                        "Medio": {"intentos":0,"partidas":0},
                        "Difícil": {"intentos":0,"partidas":0}}
    # Abro el archivo Excel para leer los datos
    workbook=load_workbook("./Estadistica.xlsx")
    hoja=workbook["Estadísticas"]
    # Recorro las filas para sumar intentos y contar partidas por dificultad
    for fila in hoja.iter_rows(min_row=2,values_only=True):
        _, _, _, dificultad, _, intentos_usados = fila
        if dificultad == 1:
            intentos_dificultad["Fácil"]["intentos"]+=intentos_usados
            intentos_dificultad["Fácil"]["partidas"]+=1
        elif dificultad == 2:
            intentos_dificultad["Medio"]["intentos"]+=intentos_usados
            intentos_dificultad["Medio"]["partidas"]+=1
        elif dificultad == 3:
            intentos_dificultad["Difícil"]["intentos"]+=intentos_usados
            intentos_dificultad["Difícil"]["partidas"]+=1
    # Calculo el promedio de intentos por dificultad
    promedios=[]
    etiquetas=[]
    for dificultad,datos in intentos_dificultad.items():
        if datos["partidas"]>0:
            promedio_intentos=datos["intentos"]/datos["partidas"]
            promedios.append(promedio_intentos)
            etiquetas.append(f"{dificultad}({promedio_intentos:.1f}intentos)")
    # Genero el gráfico circular
    plt.figure(figsize=(8,8))
    plt.pie(promedios,labels=etiquetas,autopct="%1.1f%%",startangle=140,colors=["green","orange","red"])
    plt.title("Promedio de intentos por dificultad")
    plt.show()
    return
